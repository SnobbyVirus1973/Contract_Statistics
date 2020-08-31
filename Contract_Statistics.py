import wx
import GUI
from json import dump, load
from os import remove
from os.path import splitext, exists
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, numbers, colors
from openpyxl.utils import get_column_interval


# 获取最大列数
def get_max_col(ws):
    col = 0
    none_value_count = 0
    while True:
        col += 1
        col_letter = openpyxl.utils.get_column_letter(col)
        cell = col_letter + '3'
        if not ws[cell].value:
            none_value_count += 1
        else:
            none_value_count = 0
        if none_value_count > 6:
            return col - 7


# 获取最大行数
def get_max_row(ws):
    row = 0
    none_value_count = 0
    while True:
        row += 1
        cell = f'B{row}'
        if not ws[cell].value:
            none_value_count += 1
        else:
            none_value_count = 0
        if none_value_count > 8:
            return row - 9


def delete_zero(num):
    num = str(num)
    while True:
        if len(num) == 1 or num[-1] != '0':
            return num
        else:
            num = num[:-1]
            if num[-1] == '.':
                num = num[:-1]
                return num


# 从特定格式的Excel文件中导入数据
def load_data_from_excel(wb):
    data = {'数据标题': '', '合同列表': []}
    if wb.sheetnames[0] == '资金计划表':
        ws = wb['资金计划表']
        if ws['B4'].value != '汇总' or ws['J3'].value != '备注':
            data['数据标题'] = '错误'
            return data
    else:
        data['数据标题'] = '错误'
        return data
    data['数据标题'] = ws['A1'].value
    print('最大行数为：' + str(get_max_row(ws)))
    print('最大列数为：' + str(get_max_col(ws)))
    for row in range(5, get_max_row(ws) + 1):
        print(f'当前正在识别第{row}行')
        index = -1
        name_cell_pos = 'B' + str(row)
        supplier_cell_poe = 'C' + str(row)
        signed_time_cell_poe = 'D' + str(row)
        id_cell_pos = 'E' + str(row)
        total_money_cell_pos = 'F' + str(row)
        total_paid_money_cell_pos = 'G' + str(row)
        rest_money_cell_pos = 'H' + str(row)
        paid_proportion_cell_pos = 'I' + str(row)
        comment_cell_pos = 'J' + str(row)
        if ws[id_cell_pos].value:
            index += 1
            contract = {'合同编号': ws[id_cell_pos].value if ws[id_cell_pos].value else 'xxxxxx',
                        '合同名称': ws[name_cell_pos].value if ws[name_cell_pos].value else '无合同名称',
                        '合同供应商': ws[supplier_cell_poe].value if ws[supplier_cell_poe].value else '无供应商',
                        '合同签订时间': ws[signed_time_cell_poe].value[:4] + '/' + ws[signed_time_cell_poe].value[5:7],
                        '合同总金额': delete_zero(float(ws[total_money_cell_pos].value) / 10000),
                        '合同已付金额': delete_zero(float(ws[total_paid_money_cell_pos].value) / 10000),
                        '合同剩余金额': delete_zero(float(ws[rest_money_cell_pos].value) / 10000),
                        '备注': ws[comment_cell_pos].value if ws[comment_cell_pos].value else '',
                        '合同付款明细': []
                        }
            paid_proportion = ws[paid_proportion_cell_pos].value
            paid_proportion = paid_proportion[:-1] if paid_proportion[-1] == '%' else paid_proportion
            paid_proportion = float(paid_proportion)
            contract['合同已付款比例'] = delete_zero(format(paid_proportion, '.2f')) + '%'
            data['合同列表'].append(contract)
        else:
            paid_info = {'付款说明': ws[name_cell_pos].value if ws[name_cell_pos].value else '无',
                         '付款金额': delete_zero(
                             float(ws[total_money_cell_pos].value) / 10000 if ws[total_money_cell_pos].value else 0),
                         '已付款': True if ws[total_paid_money_cell_pos].value else False
                         }
            for col in range(11, get_max_col(ws) + 1):
                col = openpyxl.utils.get_column_letter(col)
                cell_pos = str(col) + str(row)
                if ws[cell_pos].value:
                    cell_pos = str(col) + '3'
                    paid_info['付款时间'] = ws[cell_pos].value[:4] + '/' + ws[cell_pos].value[5:7]
            if not paid_info.get('付款时间'):
                paid_info['付款时间'] = data['合同列表'][index]["合同签订时间"]
            data['合同列表'][index]["合同付款明细"].append(paid_info)
    print(data)
    return data


class MainWin(GUI.MyFrame):
    # 新建数据
    def new_data(self, event):
        self.clear_value('all')
        self.data = {'数据标题': '', '合同列表': []}
        self.refresh_contract_list()
        self.main_frame_statusbar.SetLabel('新建数据完成')
        print('新建数据完成')
        # print(self.data)

    # 打开数据
    def open_data(self, event):
        self.clear_value('all')
        with wx.FileDialog(self, "打开Excel表格文件",
                           wildcard='所有支持的文件 (*.cost;*.xls;*.xlsx)|*.cost;*.xls;*.xlsx|专用数据存档 (*.cost)|*.cost|Excel '
                                    '文件 (*.xls;*.xlsx)|*.xls;*.xlsx|All Files|*',
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            pathname = fileDialog.GetPath()
            if splitext(pathname)[-1] == '.cost':
                with open(pathname, 'r', encoding='UTF-8') as f:
                    self.data = load(f)
            elif splitext(pathname)[-1] == '.xls' or splitext(pathname)[-1] == '.xlsx':
                wb = openpyxl.load_workbook(pathname, data_only=True, read_only=True)
                self.data = load_data_from_excel(wb)
                wb.close()
                if self.data['数据标题'] == '错误':
                    wx.MessageDialog(self, u"Excel文件格式不标准，无法读取", u"错误", wx.OK | wx.ICON_ERROR).ShowModal()
                    return
        self.work_title_text_ctrl.SetValue(self.data['数据标题'])
        self.refresh_contract_list()
        self.main_frame_statusbar.SetLabel('打开数据完成')
        print('打开数据完成')
        # print(self.data)

    # 保存数据
    def save_data(self, event):
        with wx.FileDialog(self, "保存合同统计数据", wildcard="专用数据存档 (*.cost)|*.cost",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            pathname = fileDialog.GetPath()
            with open(pathname, 'w', encoding='UTF-8') as file:
                dump(self.data, file, ensure_ascii=False, indent=4)
        self.main_frame_statusbar.SetLabel('保存数据完成')
        print('保存数据完成')
        # print(self.data)

    def export_data(self, event):
        self.export_to_xlsx(self.data)

    def show_help(self, event):
        wx.MessageDialog(self, u"帮助内容还未想好", u"帮助", wx.OK | wx.ICON_QUESTION).ShowModal()

    def show_info(self, event):
        wx.MessageDialog(self, u"作者：SnobbyVirus1973\n汇报bug请联系snobbyvirus1973@outlook.com\n感谢使用", u"关于",
                         wx.OK | wx.ICON_INFORMATION).ShowModal()

    # 更改数据标题时运行的函数
    def work_title_change(self, event):
        self.data['数据标题'] = self.work_title_text_ctrl.GetValue()
        self.main_frame_statusbar.SetLabel('修改数据标题完成')

    # 添加新的合同
    def add_contract(self, event):
        new_contract = {'合同名称': '新建合同', '合同编号': 'xxxxxxxx', '合同供应商': '', '合同总金额': '0', '合同已付金额': '0',
                        '合同剩余金额': '', '合同已付款比例': '0%'}
        self.data['合同列表'].append(new_contract)
        self.refresh_contract_list()
        self.main_frame_statusbar.SetLabel('添加合同完成')
        print('添加合同完成')
        # print(self.data)

    # 加载已有的合同
    def load_contract(self, event):
        self.clear_value('not_all')
        n = self.contract_list_ctrl.GetFirstSelected()
        n = self.data['合同列表'][n]
        self.contract_ID_text_ctrl.SetValue(n.get('合同编号', 'xxxxxx'))
        self.contract_title_text_ctrl.SetValue(n.get('合同名称', '无合同名称'))
        self.contract_supplier_text_ctrl.SetValue(n.get('合同供应商', '无供应商'))
        if n.get('合同签订时间'):
            self.contract_signed_year_combo_box.SetValue(n.get('合同签订时间')[:4])
            self.contract_signed_mouth_combo_box.SetValue(n.get('合同签订时间')[-2:])
        self.contract_total_money_text_ctrl.SetValue(n.get('合同总金额', '0'))
        self.contract_paid_money_text_ctrl.SetValue(n.get('合同已付金额', '0'))
        self.contract_rest_money_text_ctrl.SetValue(n.get('合同剩余金额', ''))
        self.contract_paid_proportion_text_ctrl.SetValue(n.get('合同已付款比例', '0%'))
        if n.get('合同付款明细'):
            for x in range(0, len(n.get('合同付款明细'))):
                y = x + 1
                y_checkbox = eval(f'self.pay_no{y}_checkbox')
                y_info = eval(f'self.pay_no{y}_info_text_ctrl')
                y_year = eval(f'self.pay_no{y}_year_combo_box')
                y_mouth = eval(f'self.pay_no{y}_mouth_combo_box')
                y_money = eval(f'self.pay_no{y}_money_text_ctrl')
                y_paid = eval(f'self.is_paid_no{y}_checkbox')
                y_checkbox.SetValue(1)
                y_info.SetValue(n.get('合同付款明细')[x].get('付款说明', '无'))
                y_year.SetValue(n.get('合同付款明细')[x]['付款时间'][:4])
                y_mouth.SetValue(n.get('合同付款明细')[x]['付款时间'][-2:])
                y_money.SetValue(n.get('合同付款明细')[x]['付款金额'])
                if n.get('合同付款明细')[x]['已付款']:
                    y_paid.SetValue(1)
                y_info.Enable()
                y_year.Enable()
                y_mouth.Enable()
                y_money.Enable()
                y_paid.Enable()
        self.main_frame_statusbar.SetLabel('载入合同完成')
        print('载入合同完成')
        # print(self.data)

    # 选择启用付款方式的时候执行的函数
    def pay_enable(self, event):
        n = self.FindFocus().GetLabel()[0]
        n_checkbox = eval(f'self.pay_no{n}_checkbox')
        n_info = eval(f'self.pay_no{n}_info_text_ctrl')
        n_year = eval(f'self.pay_no{n}_year_combo_box')
        n_mouth = eval(f'self.pay_no{n}_mouth_combo_box')
        n_money = eval(f'self.pay_no{n}_money_text_ctrl')
        n_paid = eval(f'self.is_paid_no{n}_checkbox')
        if n_checkbox.GetValue():
            n_info.Enable()
            n_year.Enable()
            n_mouth.Enable()
            n_money.Enable()
            n_paid.Enable()
        else:
            n_info.Disable()
            n_year.Disable()
            n_mouth.Disable()
            n_money.Disable()
            n_paid.Disable()
        self.main_frame_statusbar.SetLabel('判断付款方式是否启用完成')
        print('判断付款方式是否启用完成')
        # print(self.data)

    # 付款方式里金额变化时执行的函数
    def pay_money_changed(self, event):
        paid_money = 0
        for n in range(1, 10):
            n_checkbox = eval(f'self.pay_no{n}_checkbox.GetValue')
            n_money = eval(f'self.pay_no{n}_money_text_ctrl.GetValue')
            n_paid = eval(f'self.is_paid_no{n}_checkbox.GetValue')
            n_money = n_money()
            n_paid = n_paid()
            if n_money and n_checkbox():
                try:
                    n_money = float(n_money)
                except ValueError:
                    wx.MessageDialog(self, u"请核对是否是数字", u"错误", wx.OK | wx.ICON_ERROR).ShowModal()
                    return
                if n_paid:
                    paid_money += n_money
        self.contract_paid_money_text_ctrl.SetValue(delete_zero(paid_money))
        try:
            total_money = float(self.contract_total_money_text_ctrl.GetValue())
        except ValueError:
            wx.MessageDialog(self, u"请核对【合同总金额】是否是数字", u"错误", wx.OK | wx.ICON_ERROR).ShowModal()
            return
        self.contract_rest_money_text_ctrl.SetValue(delete_zero(total_money - paid_money))
        if total_money == 0:
            self.contract_paid_proportion_text_ctrl.SetValue('0%')
        else:
            self.contract_paid_proportion_text_ctrl.SetValue(f'{paid_money / total_money * 100}%')
        self.main_frame_statusbar.SetLabel('付款金额变动完成')
        # print('付款金额变动完成')
        # print(self.data)

    # 保存当前合同信息
    def save_contract(self, event):
        n = self.contract_list_ctrl.GetFirstSelected()
        n = self.data['合同列表'][n]
        n["合同编号"] = self.contract_ID_text_ctrl.GetValue()
        n["合同名称"] = self.contract_title_text_ctrl.GetValue()
        n["合同供应商"] = self.contract_supplier_text_ctrl.GetValue()
        n[
            "合同签订时间"] = self.contract_signed_year_combo_box.GetValue() + '/' + self.contract_signed_mouth_combo_box.GetValue()
        n["合同总金额"] = self.contract_total_money_text_ctrl.GetValue()
        n["合同已付金额"] = self.contract_paid_money_text_ctrl.GetValue()
        n["合同剩余金额"] = self.contract_rest_money_text_ctrl.GetValue()
        n["合同已付款比例"] = self.contract_paid_proportion_text_ctrl.GetValue()
        n["合同付款明细"] = []
        for i in range(1, 10):
            n_checkbox = eval(f'self.pay_no{i}_checkbox.GetValue')
            if n_checkbox():
                n_info = eval(f'self.pay_no{i}_info_text_ctrl.GetValue')
                n_year = eval(f'self.pay_no{i}_year_combo_box.GetValue')
                n_mouth = eval(f'self.pay_no{i}_mouth_combo_box.GetValue')
                n_money = eval(f'self.pay_no{i}_money_text_ctrl.GetValue')
                n_paid = eval(f'self.is_paid_no{i}_checkbox.GetValue')
                n_dir = {'付款说明': n_info(), '付款金额': n_money(), '已付款': n_paid(), '付款时间': n_year() + '/' + n_mouth()}
                n["合同付款明细"].append(n_dir)
        self.refresh_contract_list()
        self.main_frame_statusbar.SetLabel('保存合同完成')
        print('保存合同完成')
        # print(self.data)

    # 刷新合同列表
    def refresh_contract_list(self):
        self.contract_list_ctrl.DeleteAllItems()
        self.clear_value('not_all')
        print(self.data)
        for i in self.data['合同列表']:
            index = self.contract_list_ctrl.InsertItem(self.contract_list_ctrl.GetItemCount(),
                                                       str(i.get('合同编号', 'xxxxxx')))
            self.contract_list_ctrl.SetItem(index, 1, str(i.get('合同名称', '无合同名称')))
            self.contract_list_ctrl.SetItem(index, 2, str(i.get('合同供应商', '无供应商')))
            self.contract_list_ctrl.SetItem(index, 3, str(i.get('合同签订时间')))

    # 清楚右侧输入框里的数据
    def clear_value(self, argue):
        if argue == 'all':
            self.work_title_text_ctrl.Clear()
        self.contract_ID_text_ctrl.Clear()
        self.contract_title_text_ctrl.Clear()
        self.contract_supplier_text_ctrl.Clear()
        self.contract_signed_year_combo_box.SetValue('2020')
        self.contract_signed_mouth_combo_box.Select(-1)
        self.contract_total_money_text_ctrl.SetValue('0')
        self.contract_paid_money_text_ctrl.Clear()
        self.contract_rest_money_text_ctrl.Clear()
        self.contract_paid_proportion_text_ctrl.Clear()
        for n in range(1, 10):
            n_checkbox = eval(f'self.pay_no{n}_checkbox')
            n_info = eval(f'self.pay_no{n}_info_text_ctrl')
            n_year = eval(f'self.pay_no{n}_year_combo_box')
            n_mouth = eval(f'self.pay_no{n}_mouth_combo_box')
            n_money = eval(f'self.pay_no{n}_money_text_ctrl')
            n_paid = eval(f'self.is_paid_no{n}_checkbox')
            n_info.Clear()
            n_year.SetValue('2020')
            n_mouth.Select(-1)
            n_money.Clear()
            n_paid.SetValue(0)
            n_checkbox.SetValue(0)
            n_info.Disable()
            n_year.Disable()
            n_mouth.Disable()
            n_money.Disable()
            n_paid.Disable()

    # 导出为Excel文件
    def export_to_xlsx(self, export_data):
        border = Border(left=Side(border_style='medium',  # dark light
                                  color=colors.BLACK),
                        right=Side(border_style='medium',
                                   color=colors.BLACK),
                        top=Side(border_style='medium',
                                 color=colors.BLACK),
                        bottom=Side(border_style='medium',
                                    color=colors.BLACK)
                        )
        # 创建工作表
        wb = openpyxl.Workbook()
        # 创建工作簿
        funding_schedule = wb.create_sheet('资金计划表', 0)
        # 不显示网格
        funding_schedule.sheet_view.showGridLines = False
        # 新建一行输入数据标题
        line = 1
        funding_schedule.cell(line, 1).value = export_data['数据标题']
        # 新建一行输入编制、审核、批准，并添加格式
        line += 1
        funding_schedule.cell(line, 1).value = '编制：'
        funding_schedule.cell(line, 1).font = Font(name='宋体', size=11, bold=True)
        funding_schedule.cell(line, 1).number_format = numbers.FORMAT_TEXT
        funding_schedule.cell(line, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        funding_schedule.cell(line, 3).value = '审核：'
        funding_schedule.cell(line, 3).font = Font(name='宋体', size=11, bold=True)
        funding_schedule.cell(line, 3).number_format = numbers.FORMAT_TEXT
        funding_schedule.cell(line, 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        funding_schedule.cell(line, 5).value = '批准：'
        funding_schedule.cell(line, 5).font = Font(name='宋体', size=11, bold=True)
        funding_schedule.cell(line, 5).number_format = numbers.FORMAT_TEXT
        funding_schedule.cell(line, 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 新建一行输入标题行
        line += 1
        funding_schedule.cell(line, 1).value = '序号'
        funding_schedule.cell(line, 2).value = '合同名称'
        funding_schedule.cell(line, 3).value = '供应商'
        funding_schedule.cell(line, 4).value = '合同签订时间'
        funding_schedule.cell(line, 5).value = '合同编号'
        funding_schedule.cell(line, 6).value = '总金额（元）'
        funding_schedule.cell(line, 7).value = '累计已付款（元）'
        funding_schedule.cell(line, 8).value = '余款总额（元）'
        funding_schedule.cell(line, 9).value = '累计已付款比例'
        funding_schedule.cell(line, 10).value = '备注'
        # 统计每个合同的付款方式的月份，并添加到标题栏
        time_list = []
        for x in export_data['合同列表']:
            for y in x['合同付款明细']:
                z = y['付款时间'][:4] + y['付款时间'][-2:]
                time_list.append(z)
        # 每个合同的付款方式的月份进行排序和去重
        time_list.sort()
        time_list_new = []
        for i in time_list:
            if i not in time_list_new:
                time_list_new.append(i)
        time_list = time_list_new
        # 将每个合同的付款方式的月份添加到标题栏
        for index, time in enumerate(time_list):
            funding_schedule.cell(line, 11 + index).value = f'{time[:4]}年{time[-2:]}月付款'
        # 计算最大列数
        max_col = get_max_col(funding_schedule)
        # 设置第三行标题行格式
        for col in range(1, max_col + 1):
            funding_schedule.cell(line, col).font = Font(name='宋体', size=10, bold=True)
            funding_schedule.cell(line, col).number_format = numbers.FORMAT_TEXT
            funding_schedule.cell(line, col).border = border
            funding_schedule.cell(line, col).alignment = Alignment(horizontal='center', vertical='center',
                                                                   wrap_text=True)
        # 合并标题的单元格
        funding_schedule.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        # 设置第一行标题行的格式
        funding_schedule.cell(1, 1).font = Font(name='宋体', size=18, bold=True)
        funding_schedule.cell(1, 1).number_format = numbers.FORMAT_TEXT
        funding_schedule.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # 先建立汇总行，后面再计算金额总数
        line += 1
        funding_schedule.cell(line, 2).value = '汇总'
        # 读取每个合同，将信息添加进去
        for index, contract in enumerate(export_data['合同列表']):
            line += 1
            funding_schedule.cell(line, 1).value = f'{index + 1}'
            funding_schedule.cell(line, 2).value = contract['合同名称']
            funding_schedule.cell(line, 3).value = contract['合同供应商']
            funding_schedule.cell(line, 4).value = contract['合同签订时间']
            funding_schedule.cell(line, 5).value = contract['合同编号']
            funding_schedule.cell(line, 6).value = delete_zero(float(contract['合同总金额']) * 10000)
            funding_schedule.cell(line, 7).value = delete_zero(float(contract['合同已付金额']) * 10000)
            funding_schedule.cell(line, 8).value = delete_zero(float(contract['合同剩余金额']) * 10000)
            funding_schedule.cell(line, 9).value = contract['合同已付款比例']
            funding_schedule.cell(line, 10).value = contract.get('备注', '')
            # 添加格式
            for col in range(1, max_col + 1):
                funding_schedule.cell(line, col).font = Font(name='宋体', size=9, bold=False)
                funding_schedule.cell(line, col).number_format = numbers.FORMAT_TEXT
                funding_schedule.cell(line, col).border = border
                funding_schedule.cell(line, col).fill = PatternFill(fill_type='solid', fgColor='FFC125')
                if col == 10:
                    funding_schedule.cell(line, col).alignment = Alignment(horizontal='left', vertical='center',
                                                                           wrap_text=True)
                else:
                    funding_schedule.cell(line, col).alignment = Alignment(horizontal='center', vertical='center',
                                                                           wrap_text=True)
            # 此时line存储的是合同信息的行数，new_line用来存储每一个付款方式的行数
            new_line = 0
            for index_, paid_info in enumerate(contract['合同付款明细']):
                new_line = line + index_ + 1
                funding_schedule.cell(new_line, 2).value = paid_info['付款说明']
                funding_schedule.cell(new_line, 6).value = delete_zero(float(paid_info['付款金额']) * 10000)
                if paid_info['已付款']:
                    funding_schedule.cell(new_line, 7).value = delete_zero(float(paid_info['付款金额']) * 10000)
                    funding_schedule.cell(new_line, 8).value = '0'
                else:
                    funding_schedule.cell(new_line, 7).value = '0'
                    funding_schedule.cell(new_line, 8).value = delete_zero(float(paid_info['付款金额']) * 10000)
                paid_time = paid_info['付款时间'][:4] + paid_info['付款时间'][-2:]
                # 定位付款时间的列数
                paid_index = time_list.index(paid_time)
                funding_schedule.cell(line, 11 + paid_index).value = delete_zero(float(paid_info['付款金额']) * 10000)
                funding_schedule.cell(new_line, 11 + paid_index).value = delete_zero(float(paid_info['付款金额']) * 10000)
                # 添加格式
                for col in range(1, max_col + 1):
                    funding_schedule.cell(new_line, col).font = Font(name='宋体', size=9, bold=False)
                    funding_schedule.cell(new_line, col).number_format = numbers.FORMAT_TEXT
                    funding_schedule.cell(new_line, col).border = border
                    if col == 2 or col == 10:
                        funding_schedule.cell(new_line, col).alignment = Alignment(horizontal='left', vertical='center',
                                                                                   wrap_text=True)
                    else:
                        funding_schedule.cell(new_line, col).alignment = Alignment(horizontal='center',
                                                                                   vertical='center', wrap_text=True)
            # 折叠付款明细的每行
            funding_schedule.row_dimensions.group(line + 1, new_line, hidden=True)
            line = new_line

        # 设置行高
        funding_schedule.row_dimensions[1].height = 43.5
        for i in range(2, get_max_row(funding_schedule) + 1):
            funding_schedule.row_dimensions[i].height = 28.5
        # 设置列宽
        funding_schedule.column_dimensions['A'].width = 4.88
        funding_schedule.column_dimensions['B'].width = 45.13
        funding_schedule.column_dimensions['C'].width = 31.13
        funding_schedule.column_dimensions['D'].width = 12.13
        funding_schedule.column_dimensions['E'].width = 15.75
        funding_schedule.column_dimensions['F'].width = 18.38
        funding_schedule.column_dimensions['G'].width = 18.38
        funding_schedule.column_dimensions['H'].width = 18.38
        funding_schedule.column_dimensions['I'].width = 18.38
        funding_schedule.column_dimensions['j'].width = 18.38
        for i in range(11, max_col + 1):
            funding_schedule.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 26.38
        # 设置打印格式
        funding_schedule.print_options.horizontalCentered = True
        funding_schedule.print_options.verticalCentered = True
        # 设置每页均打印前三行
        funding_schedule.print_title_rows = '1:3'
        with wx.FileDialog(self, "导出合同统计数据", wildcard="Excel (*.xlsx)|*.xlsx",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            pathname = fileDialog.GetPath()
            if exists(pathname):
                try:
                    remove(pathname)
                except PermissionError:
                    wx.MessageDialog(self, u"另一个程序正在使用此文件，进程无法访问。请重新导出", u"错误", wx.OK | wx.ICON_ERROR).ShowModal()
                    return
            wb.save(pathname)
        self.main_frame_statusbar.SetLabel('导出数据完成')
        print('导出数据完成')


if __name__ == '__main__':
    app = wx.App()
    frame = MainWin(None, wx.ID_ANY, "")
    app.SetTopWindow(frame)
    frame.Show()
    app.MainLoop()
