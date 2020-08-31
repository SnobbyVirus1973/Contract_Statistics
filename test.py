import os
import openpyxl
from openpyxl.utils import get_column_interval
''''
i = {1:10, 2:20}
print(i.keys())
print( 1 == 1.00)
for i in range(0,9):
    print(i)
'''

'''
time_list = ['202001', '202005', '202005', '202004', '202006', '202004', '202008', '202010', '202012', '202011', '202001',
             '202101', '202102', '202001', '202006', '202006', '202005', '202107', '202012', '202012', '202011', '202001']
time_list.sort()
time_list_new = []
for i in time_list:
    if i not in time_list_new:
        time_list_new.append(i)
time_list = time_list_new
print(time_list)
print(time_list.index('202101'))

print(os.path.exists())
'''

print(openpyxl.utils.get_column_letter(10))


'''
wb = openpyxl.Workbook()
ws = wb.create_sheet('资金计划表', 0)
ws['A1'].
'''